import json

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from excel_mcp.calculations import inspect_formula
from excel_mcp.server import inspect_formula as inspect_formula_tool
from excel_mcp.server import detect_circular_dependencies as detect_circular_dependencies_tool
from excel_mcp.workbook import detect_circular_dependencies


def _load_tool_payload(raw: str) -> dict:
    payload = json.loads(raw)
    assert payload["ok"] is True
    assert "operation" in payload
    assert "message" in payload
    return payload


def test_inspect_formula_reports_functions_and_reference_types():
    result = inspect_formula("SUM(A1:B2)+ReportTotal+SalesTable[Amount]+TODAY()")

    assert result["formula"] == "=SUM(A1:B2)+ReportTotal+SalesTable[Amount]+TODAY()"
    assert result["syntax_valid"] is True
    assert result["summary"]["function_count"] == 2
    assert result["summary"]["reference_count"] == 3
    assert result["summary"]["uses_volatile_functions"] is True
    assert result["summary"]["uses_unsafe_functions"] is False
    assert [item["name"] for item in result["functions"]] == ["SUM", "TODAY"]
    assert [item["reference_type"] for item in result["references"]] == [
        "worksheet_range",
        "named_or_identifier",
        "structured_reference",
    ]


def test_inspect_formula_reports_unsafe_functions_without_failing():
    result = inspect_formula("INDIRECT(\"A1\")")

    assert result["syntax_valid"] is True
    assert result["summary"]["uses_unsafe_functions"] is True
    assert result["summary"]["uses_volatile_functions"] is True
    assert result["functions"][0]["name"] == "INDIRECT"
    assert result["functions"][0]["unsafe"] is True


def test_inspect_formula_tool_returns_envelope():
    payload = _load_tool_payload(inspect_formula_tool("SUM(A1:A5)"))

    assert payload["operation"] == "inspect_formula"
    assert payload["data"]["formula"] == "=SUM(A1:A5)"
    assert payload["data"]["summary"]["function_count"] == 1


def test_detect_circular_dependencies_reports_multi_cell_cycle(tmp_path):
    filepath = str(tmp_path / "cycle.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=B1+1"
    ws["B1"] = "=A1+1"
    wb.save(filepath)
    wb.close()

    result = detect_circular_dependencies(filepath)

    assert result["summary"]["has_circular_dependencies"] is True
    assert result["summary"]["cycle_count"] == 1
    assert result["summary"]["multi_cell_cycle_count"] == 1
    assert result["summary"]["self_referential_cycle_count"] == 0
    cycle = result["cycles"]["sample"][0]
    assert cycle["size"] == 2
    assert {item["cell"] for item in cycle["cells"]} == {"A1", "B1"}


def test_detect_circular_dependencies_reports_self_reference(tmp_path):
    filepath = str(tmp_path / "self-cycle.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=A1+1"
    wb.save(filepath)
    wb.close()

    result = detect_circular_dependencies(filepath)

    assert result["summary"]["has_circular_dependencies"] is True
    assert result["summary"]["cycle_count"] == 1
    assert result["summary"]["self_referential_cycle_count"] == 1
    cycle = result["cycles"]["sample"][0]
    assert cycle["includes_self_reference"] is True
    assert cycle["cells"][0]["cell"] == "A1"


def test_detect_circular_dependencies_tracks_named_range_cycles(tmp_path):
    filepath = str(tmp_path / "named-cycle.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=B1+1"
    ws["B1"] = "=SUM(LoopRange)"
    wb.defined_names["LoopRange"] = DefinedName("LoopRange", attr_text="Sheet1!$A$1")
    wb.save(filepath)
    wb.close()

    result = detect_circular_dependencies(filepath)

    assert result["summary"]["cycle_count"] == 1
    cycle = result["cycles"]["sample"][0]
    assert {item["cell"] for item in cycle["cells"]} == {"A1", "B1"}


def test_detect_circular_dependencies_tool_returns_envelope(tmp_path):
    filepath = str(tmp_path / "tool-cycle.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=B1"
    ws["B1"] = "=A1"
    wb.save(filepath)
    wb.close()

    payload = _load_tool_payload(detect_circular_dependencies_tool(filepath))

    assert payload["operation"] == "detect_circular_dependencies"
    assert payload["data"]["summary"]["cycle_count"] == 1
