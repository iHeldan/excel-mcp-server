import pytest
from openpyxl import load_workbook

from excel_mcp.formatting import format_range
from excel_mcp.exceptions import ValidationError, FormattingError


# --- Basic formatting ---

def test_bold_formatting(tmp_workbook):
    result = format_range(tmp_workbook, "Sheet1", "A1", bold=True)
    assert "Applied" in result["message"]

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].font.bold is True
    wb.close()


def test_italic_and_underline(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", italic=True, underline=True)

    wb = load_workbook(tmp_workbook)
    cell = wb["Sheet1"]["A1"]
    assert cell.font.italic is True
    assert cell.font.underline == "single"
    wb.close()


def test_font_size(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", font_size=16)

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].font.size == 16
    wb.close()


def test_font_color(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", font_color="FF0000")

    wb = load_workbook(tmp_workbook)
    # openpyxl may store as FFFF0000 or 00FF0000 depending on version
    assert wb["Sheet1"]["A1"].font.color.rgb in ("FFFF0000", "00FF0000")
    wb.close()


def test_background_color(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", bg_color="00FF00")

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].fill.start_color.rgb == "FF00FF00"
    wb.close()


# --- Range formatting ---

def test_format_range_multiple_cells(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", end_cell="C1", bold=True, font_size=14)

    wb = load_workbook(tmp_workbook)
    for col in ["A", "B", "C"]:
        cell = wb["Sheet1"][f"{col}1"]
        assert cell.font.bold is True
        assert cell.font.size == 14
    wb.close()


# --- Border ---

def test_border_formatting(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", border_style="thin")

    wb = load_workbook(tmp_workbook)
    border = wb["Sheet1"]["A1"].border
    assert border.left.style == "thin"
    assert border.right.style == "thin"
    wb.close()


# --- Alignment & wrap ---

def test_alignment_center(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", alignment="center")

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].alignment.horizontal == "center"
    wb.close()


def test_wrap_text(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", wrap_text=True)

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["A1"].alignment.wrap_text is True
    wb.close()


# --- Number format ---

def test_number_format(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "B2", number_format="#,##0.00")

    wb = load_workbook(tmp_workbook)
    assert wb["Sheet1"]["B2"].number_format == "#,##0.00"
    wb.close()


# --- Merge cells ---

def test_merge_cells(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", end_cell="C1", merge_cells=True)

    wb = load_workbook(tmp_workbook)
    merged = list(wb["Sheet1"].merged_cells.ranges)
    assert any("A1:C1" in str(r) for r in merged)
    wb.close()


# --- Dry run ---

def test_dry_run_does_not_persist(tmp_workbook):
    format_range(tmp_workbook, "Sheet1", "A1", bold=True, dry_run=True)

    wb = load_workbook(tmp_workbook)
    # dry_run should NOT save changes
    assert wb["Sheet1"]["A1"].font.bold is not True
    wb.close()


# --- Conditional formatting ---

def test_conditional_format_cell_is(tmp_workbook):
    cond = {
        "type": "cell_is",
        "params": {
            "operator": "greaterThan",
            "formula": ["30"],
            "fill": {"fgColor": "FFC7CE"},
        },
    }
    result = format_range(
        tmp_workbook, "Sheet1", "B2", end_cell="B6", conditional_format=cond
    )
    assert "Applied" in result["message"]


def test_conditional_format_missing_type(tmp_workbook):
    with pytest.raises(FormattingError, match="type not specified"):
        format_range(
            tmp_workbook, "Sheet1", "A1", conditional_format={"params": {}}
        )


# --- Error cases ---

def test_format_invalid_sheet(tmp_workbook):
    with pytest.raises(ValidationError, match="not found"):
        format_range(tmp_workbook, "NoSheet", "A1", bold=True)


def test_format_invalid_start_cell(tmp_workbook):
    with pytest.raises(ValidationError, match="Invalid start cell"):
        format_range(tmp_workbook, "Sheet1", "123", bold=True)
