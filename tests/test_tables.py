import pytest
from openpyxl import load_workbook

from excel_mcp.tables import create_excel_table
from excel_mcp.exceptions import DataError


def test_create_table_with_auto_name(tmp_workbook):
    result = create_excel_table(tmp_workbook, "Sheet1", "A1:C6")
    assert "Successfully created table" in result["message"]
    assert result["range"] == "A1:C6"
    assert result["table_name"].startswith("Table_")


def test_create_table_with_custom_name(tmp_workbook):
    result = create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="MyTable")
    assert result["table_name"] == "MyTable"

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    table_names = [t.displayName for t in ws.tables.values()]
    assert "MyTable" in table_names
    wb.close()


def test_create_table_with_custom_style(tmp_workbook):
    result = create_excel_table(
        tmp_workbook, "Sheet1", "A1:C6", table_style="TableStyleLight1"
    )
    assert "Successfully created table" in result["message"]


def test_create_table_persists_to_file(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C6", table_name="PersistTest")

    wb = load_workbook(tmp_workbook)
    ws = wb["Sheet1"]
    assert len(ws.tables) == 1
    wb.close()


def test_create_table_sheet_not_found(tmp_workbook):
    with pytest.raises(DataError, match="not found"):
        create_excel_table(tmp_workbook, "NoSheet", "A1:C6")


def test_create_table_duplicate_name(tmp_workbook):
    create_excel_table(tmp_workbook, "Sheet1", "A1:C3", table_name="Dupl")
    with pytest.raises(DataError):
        create_excel_table(tmp_workbook, "Sheet1", "A1:C3", table_name="Dupl")
