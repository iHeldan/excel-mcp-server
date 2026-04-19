import json
import shutil

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from excel_mcp.server import (
    analyze_range_impact,
    apply_workbook_repairs,
    audit_workbook,
    bulk_aggregate_workbooks,
    create_chart,
    create_table,
    create_workbook,
    create_worksheet,
    describe_sheet_layout,
    diff_workbooks,
    find_free_canvas,
    list_all_sheets,
    plan_workbook_repairs,
    profile_workbook,
    write_data_to_excel,
)


def _load_tool_payload(raw: str) -> dict:
    payload = json.loads(raw)
    assert payload["ok"] is True
    assert "operation" in payload
    return payload


def test_agent_workflow_orientation_mutation_and_diff(tmp_path):
    filepath = str(tmp_path / "agent-workflow.xlsx")
    before_path = str(tmp_path / "agent-workflow-before.xlsx")

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    data_sheet.append(["Month", "Sales"])
    for row in [("Jan", 120), ("Feb", 140), ("Mar", 160), ("Apr", 180)]:
        data_sheet.append(list(row))

    chart = BarChart()
    data = Reference(data_sheet, min_col=2, min_row=1, max_row=5)
    categories = Reference(data_sheet, min_col=1, min_row=2, max_row=5)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = "Sales"
    data_sheet.add_chart(chart, "E2")

    dashboard = workbook.create_sheet("Dashboard")
    dashboard["A1"] = "Executive View"
    dashboard["A3"] = "=SUM(Data!B2:B5)"
    workbook.save(filepath)
    workbook.close()

    _load_tool_payload(create_table(filepath, "Data", "A1:B5", "SalesTable"))
    shutil.copyfile(filepath, before_path)

    sheets_payload = _load_tool_payload(list_all_sheets(filepath))
    assert [sheet["name"] for sheet in sheets_payload["data"]["sheets"]] == ["Data", "Dashboard"]

    profile_payload = _load_tool_payload(profile_workbook(filepath))
    assert profile_payload["data"]["sheet_count"] == 2
    assert profile_payload["data"]["table_count"] == 1
    assert profile_payload["data"]["chart_count"] == 1

    layout_payload = _load_tool_payload(describe_sheet_layout(filepath, "Dashboard"))
    assert layout_payload["data"]["sheet_name"] == "Dashboard"
    assert layout_payload["data"]["summary"]["chart_count"] == 0

    impact_payload = _load_tool_payload(analyze_range_impact(filepath, "Data", "B2:B5"))
    assert impact_payload["data"]["summary"]["risk_level"] in {"medium", "high"}
    assert impact_payload["data"]["summary"]["dependent_formula_count"] >= 1

    write_payload = _load_tool_payload(
        write_data_to_excel(
            filepath,
            "Data",
            [["Jan", 125], ["Feb", 150], ["Mar", 175], ["Apr", 210]],
            "A2",
        )
    )
    assert write_payload["data"]["changed_cells"] == 4

    diff_payload = _load_tool_payload(diff_workbooks(before_path, filepath))
    assert diff_payload["data"]["cell_changes"]["count"] == 4
    changed_cells = {item["cell"] for item in diff_payload["data"]["cell_changes"]["sample"]}
    assert changed_cells == {"B2", "B3", "B4", "B5"}


def test_agent_workflow_repair_preview_apply_and_verify(tmp_path):
    filepath = str(tmp_path / "repair-workflow.xlsx")

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Data"
    ws.append(["Choice", "Value"])
    ws.append(["A", 10])
    ws.append(["B", 20])

    validation = DataValidation(type="list", formula1="=MissingSheet!$A$1:$A$3")
    validation.add("A2:A10")
    ws.add_data_validation(validation)

    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws.conditional_formatting.add(
        "B2:B10",
        FormulaRule(formula=["MissingSheet!$A$1>0"], fill=fill),
    )
    workbook.defined_names["BrokenName"] = DefinedName(
        "BrokenName",
        attr_text="MissingSheet!$A$1",
    )
    workbook.save(filepath)
    workbook.close()

    audit_payload = _load_tool_payload(audit_workbook(filepath))
    assert audit_payload["data"]["summary"]["risk_level"] == "high"
    assert audit_payload["data"]["summary"]["high_count"] >= 1

    plan_payload = _load_tool_payload(plan_workbook_repairs(filepath))
    assert plan_payload["data"]["step_count"] >= 3
    assert len(plan_payload["data"]["quick_wins"]) >= 1

    preview_payload = _load_tool_payload(apply_workbook_repairs(filepath, dry_run=True))
    assert preview_payload["dry_run"] is True
    assert preview_payload["data"]["action_count"] == 3
    assert preview_payload["data"]["diff"]["summary"]["named_range_change_count"] == 0

    apply_payload = _load_tool_payload(apply_workbook_repairs(filepath, dry_run=False))
    assert apply_payload["dry_run"] is False
    assert apply_payload["data"]["action_count"] == 3
    assert apply_payload["data"]["audit_after"]["risk_level"] == "low"

    after_payload = _load_tool_payload(audit_workbook(filepath))
    assert after_payload["data"]["summary"]["risk_level"] == "low"
    assert after_payload["data"]["summary"]["high_count"] == 0


def test_agent_workflow_multi_workbook_report_to_dashboard_tab(tmp_path):
    north = str(tmp_path / "north.xlsx")
    south = str(tmp_path / "south.xlsx")
    report = str(tmp_path / "report.xlsx")

    for filepath, rows in [
        (north, [("Alpha", "North", 1200), ("Beta", "North", 900)]),
        (south, [("Alpha", "South", 1000), ("Gamma", "South", 700)]),
    ]:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Sales"
        worksheet.append(["Campaign", "Region", "Revenue"])
        for row in rows:
            worksheet.append(list(row))
        workbook.save(filepath)
        workbook.close()
        _load_tool_payload(create_table(filepath, "Sales", "A1:C3", "SalesTable"))

    _load_tool_payload(create_workbook(report))
    _load_tool_payload(create_worksheet(report, "Summary"))

    aggregate_payload = _load_tool_payload(
        bulk_aggregate_workbooks(
            [north, south],
            table_name="SalesTable",
            group_by=["Region"],
            metrics=[{"agg": "sum", "column": "Revenue", "as": "Total Revenue"}],
        )
    )
    assert aggregate_payload["data"]["rows"] == [
        ["North", 2100],
        ["South", 1700],
    ]

    rows = [aggregate_payload["data"]["headers"]] + aggregate_payload["data"]["rows"]
    _load_tool_payload(write_data_to_excel(report, "Summary", rows, "A1"))

    canvas_payload = _load_tool_payload(find_free_canvas(report, "Summary", min_rows=8, min_cols=6, limit=1))
    suggestion = canvas_payload["data"]["suggestions"][0]
    assert suggestion["anchor_cell"]

    chart_payload = _load_tool_payload(
        create_chart(
            report,
            "Summary",
            "bar",
            data_range="A1:B3",
            target_cell=suggestion["anchor_cell"],
            title="Revenue by Region",
            x_axis="Region",
            y_axis="Revenue",
        )
    )
    assert chart_payload["data"]["details"]["type"] == "bar"

    layout_payload = _load_tool_payload(describe_sheet_layout(report, "Summary"))
    assert layout_payload["data"]["summary"]["chart_count"] == 1
    assert layout_payload["data"]["used_range"] == "A1:B3"
