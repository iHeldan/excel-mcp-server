import pytest
from openpyxl import load_workbook

from excel_mcp.chart import ChartType, create_chart_in_sheet
from excel_mcp.exceptions import ValidationError, ChartError


@pytest.fixture
def chart_workbook(tmp_path):
    """Workbook with numeric data suitable for charting."""
    from openpyxl import Workbook

    filepath = str(tmp_path / "chart.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws["A1"] = "Month"
    ws["B1"] = "Revenue"
    ws["C1"] = "Cost"
    for i, (month, rev, cost) in enumerate(
        [("Jan", 100, 60), ("Feb", 150, 80), ("Mar", 200, 90), ("Apr", 180, 85)],
        start=2,
    ):
        ws[f"A{i}"] = month
        ws[f"B{i}"] = rev
        ws[f"C{i}"] = cost
    wb.save(filepath)
    wb.close()
    return filepath


# --- ChartType enum ---

def test_chart_type_enum_has_five_members():
    assert len(ChartType) == 5
    assert set(ChartType) == {
        ChartType.LINE,
        ChartType.BAR,
        ChartType.PIE,
        ChartType.SCATTER,
        ChartType.AREA,
    }


# --- Successful chart creation ---

@pytest.mark.parametrize("chart_type", ["line", "bar", "pie", "area"])
def test_create_chart_supported_types(chart_workbook, chart_type):
    result = create_chart_in_sheet(
        chart_workbook, "Sales", "A1:B5", chart_type, "E1", title=f"Test {chart_type}"
    )
    assert "successfully" in result["message"].lower()
    assert result["details"]["type"] == chart_type


def test_create_scatter_chart(chart_workbook):
    result = create_chart_in_sheet(
        chart_workbook, "Sales", "B1:C5", "scatter", "E1", title="Scatter"
    )
    assert result["details"]["type"] == "scatter"


def test_chart_with_style_options(chart_workbook):
    style = {
        "show_legend": True,
        "legend_position": "b",
        "show_data_labels": True,
        "data_label_options": {"show_val": True, "show_percent": False},
        "grid_lines": True,
    }
    result = create_chart_in_sheet(
        chart_workbook, "Sales", "A1:B5", "bar", "E1", title="Styled", style=style
    )
    assert "successfully" in result["message"].lower()


def test_chart_without_legend(chart_workbook):
    result = create_chart_in_sheet(
        chart_workbook, "Sales", "A1:B5", "line", "E1",
        style={"show_legend": False, "show_data_labels": False},
    )
    assert "successfully" in result["message"].lower()


def test_chart_with_axis_labels(chart_workbook):
    result = create_chart_in_sheet(
        chart_workbook, "Sales", "A1:B5", "bar", "E1",
        title="Revenue", x_axis="Month", y_axis="EUR",
    )
    assert result["details"]["data_range"] == "A1:B5"


# --- Error cases ---

def test_chart_invalid_sheet(chart_workbook):
    with pytest.raises(ValidationError, match="not found"):
        create_chart_in_sheet(chart_workbook, "NoSheet", "A1:B5", "bar", "E1")


def test_chart_unsupported_type(chart_workbook):
    with pytest.raises(ValidationError, match="Unsupported chart type"):
        create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "radar", "E1")


def test_chart_invalid_data_range(chart_workbook):
    with pytest.raises(ValidationError, match="Invalid data range"):
        create_chart_in_sheet(chart_workbook, "Sales", "ZZZ", "bar", "E1")


def test_chart_invalid_target_cell(chart_workbook):
    with pytest.raises((ValidationError, ChartError)):
        create_chart_in_sheet(chart_workbook, "Sales", "A1:B5", "bar", "")


def test_chart_cross_sheet_reference_invalid(chart_workbook):
    with pytest.raises(ValidationError, match="not found"):
        create_chart_in_sheet(chart_workbook, "Sales", "Missing!A1:B5", "bar", "E1")
