import pytest
from openpyxl import load_workbook

from excel_mcp.pivot import (
    create_pivot_table,
    _aggregate_values,
    _filter_data,
    _get_combinations,
)
from excel_mcp.exceptions import ValidationError, PivotError


@pytest.fixture
def pivot_workbook(tmp_path):
    """Workbook with categorical + numeric data for pivot testing."""
    from openpyxl import Workbook

    filepath = str(tmp_path / "pivot.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Region"
    ws["B1"] = "Product"
    ws["C1"] = "Sales"
    ws["D1"] = "Quantity"
    rows = [
        ("North", "Widget", 100, 10),
        ("North", "Gadget", 200, 5),
        ("South", "Widget", 150, 8),
        ("South", "Gadget", 300, 12),
        ("North", "Widget", 120, 6),
    ]
    for i, (region, product, sales, qty) in enumerate(rows, start=2):
        ws[f"A{i}"] = region
        ws[f"B{i}"] = product
        ws[f"C{i}"] = sales
        ws[f"D{i}"] = qty
    wb.save(filepath)
    wb.close()
    return filepath


# --- Successful pivot creation ---

def test_pivot_default_sum(pivot_workbook):
    result = create_pivot_table(
        pivot_workbook, "Data", "A1:D6", rows=["Region"], values=["Sales"]
    )
    assert result["details"]["aggregation"] == "sum"
    assert result["details"]["pivot_sheet"] == "Data_pivot"

    wb = load_workbook(pivot_workbook)
    pivot_ws = wb["Data_pivot"]
    assert pivot_ws["A1"].value == "Region"
    assert "Sales (sum)" in pivot_ws["B1"].value
    wb.close()


def test_pivot_average_aggregation(pivot_workbook):
    result = create_pivot_table(
        pivot_workbook, "Data", "A1:D6", rows=["Region"], values=["Sales"], agg_func="average"
    )
    assert result["details"]["aggregation"] == "average"

    wb = load_workbook(pivot_workbook)
    pivot_ws = wb["Data_pivot"]
    # North: (100+200+120)/3 = 140, South: (150+300)/2 = 225
    north_val = pivot_ws["B2"].value
    south_val = pivot_ws["B3"].value
    assert abs(north_val - 140.0) < 0.01
    assert abs(south_val - 225.0) < 0.01
    wb.close()


def test_pivot_multiple_row_fields(pivot_workbook):
    create_pivot_table(pivot_workbook, "Data", "A1:D6", rows=["Region", "Product"], values=["Sales"])
    wb = load_workbook(pivot_workbook)
    pivot_ws = wb["Data_pivot"]
    assert pivot_ws["A1"].value == "Region"
    assert pivot_ws["B1"].value == "Product"
    wb.close()


def test_pivot_multiple_value_fields(pivot_workbook):
    create_pivot_table(pivot_workbook, "Data", "A1:D6", rows=["Region"], values=["Sales", "Quantity"])
    wb = load_workbook(pivot_workbook)
    pivot_ws = wb["Data_pivot"]
    assert "Sales (sum)" in pivot_ws["B1"].value
    assert "Quantity (sum)" in pivot_ws["C1"].value
    wb.close()


def test_pivot_count_aggregation(pivot_workbook):
    create_pivot_table(pivot_workbook, "Data", "A1:D6", rows=["Region"], values=["Sales"], agg_func="count")
    wb = load_workbook(pivot_workbook)
    pivot_ws = wb["Data_pivot"]
    # North has 3 rows, South has 2
    vals = {pivot_ws[f"A{r}"].value: pivot_ws[f"B{r}"].value for r in range(2, 4)}
    assert vals["North"] == 3
    assert vals["South"] == 2
    wb.close()


def test_pivot_replaces_existing_pivot_sheet(pivot_workbook):
    create_pivot_table(pivot_workbook, "Data", "A1:D6", rows=["Region"], values=["Sales"])
    # Run again — should replace, not fail
    result = create_pivot_table(
        pivot_workbook, "Data", "A1:D6", rows=["Region"], values=["Sales"], agg_func="max"
    )
    assert result["details"]["aggregation"] == "max"


def test_pivot_resolves_case_insensitive_field_names(pivot_workbook):
    result = create_pivot_table(
        pivot_workbook, "Data", "A1:D6", rows=["region"], values=["sales"]
    )

    assert result["details"]["rows"] == ["Region"]
    assert result["details"]["values"] == ["Sales"]

    wb = load_workbook(pivot_workbook)
    pivot_ws = wb["Data_pivot"]
    vals = {pivot_ws[f"A{r}"].value: pivot_ws[f"B{r}"].value for r in range(2, 4)}
    assert vals["North"] == 420
    assert vals["South"] == 450
    wb.close()


def test_pivot_columns_create_grouped_value_columns(pivot_workbook):
    result = create_pivot_table(
        pivot_workbook,
        "Data",
        "A1:D6",
        rows=["Region"],
        values=["Sales"],
        columns=["Product"],
    )

    assert result["details"]["columns"] == ["Product"]

    wb = load_workbook(pivot_workbook)
    pivot_ws = wb["Data_pivot"]
    headers = [pivot_ws["A1"].value, pivot_ws["B1"].value, pivot_ws["C1"].value]
    assert headers == ["Region", "Gadget", "Widget"]
    north = [pivot_ws["A2"].value, pivot_ws["B2"].value, pivot_ws["C2"].value]
    south = [pivot_ws["A3"].value, pivot_ws["B3"].value, pivot_ws["C3"].value]
    assert north == ["North", 200, 220]
    assert south == ["South", 300, 150]
    wb.close()


# --- Error cases ---

def test_pivot_invalid_agg_func(pivot_workbook):
    with pytest.raises(ValidationError, match="Invalid aggregation function"):
        create_pivot_table(
            pivot_workbook, "Data", "A1:D6", rows=["Region"], values=["Sales"], agg_func="mean"
        )


def test_pivot_invalid_row_field(pivot_workbook):
    with pytest.raises(ValidationError, match="Invalid row field"):
        create_pivot_table(
            pivot_workbook, "Data", "A1:D6", rows=["Missing"], values=["Sales"]
        )


def test_pivot_invalid_value_field(pivot_workbook):
    with pytest.raises(ValidationError, match="Invalid value field"):
        create_pivot_table(
            pivot_workbook, "Data", "A1:D6", rows=["Region"], values=["Nonexistent"]
        )


def test_pivot_invalid_range_format(pivot_workbook):
    with pytest.raises(ValidationError, match="Data range must be in format"):
        create_pivot_table(pivot_workbook, "Data", "A1", rows=["Region"], values=["Sales"])


def test_pivot_sheet_not_found(pivot_workbook):
    with pytest.raises(PivotError, match="not found"):
        create_pivot_table(
            pivot_workbook, "NoSheet", "A1:D6", rows=["Region"], values=["Sales"]
        )


# --- Unit tests for helper functions ---

def test_aggregate_sum():
    data = [{"val": 10}, {"val": 20}, {"val": 30}]
    assert _aggregate_values(data, "val", "sum") == 60


def test_aggregate_average():
    data = [{"val": 10}, {"val": 20}, {"val": 30}]
    assert _aggregate_values(data, "val", "average") == 20.0


def test_aggregate_count():
    data = [{"val": 10}, {"val": 20}]
    assert _aggregate_values(data, "val", "count") == 2


def test_aggregate_min_max():
    data = [{"val": 5}, {"val": 15}, {"val": 10}]
    assert _aggregate_values(data, "val", "min") == 5
    assert _aggregate_values(data, "val", "max") == 15


def test_aggregate_empty_returns_zero():
    assert _aggregate_values([], "val", "sum") == 0


def test_filter_data():
    data = [
        {"region": "North", "val": 10},
        {"region": "South", "val": 20},
        {"region": "North", "val": 30},
    ]
    filtered = _filter_data(data, {"region": "North"}, {})
    assert len(filtered) == 2
    assert all(r["region"] == "North" for r in filtered)


def test_get_combinations():
    field_values = {"a": ["x", "y"], "b": ["1", "2"]}
    combos = _get_combinations(field_values)
    assert len(combos) == 4
    assert {"a": "x", "b": "1"} in combos
