from enum import Enum
import logging
import math
from typing import Any, Dict, List, Optional

from openpyxl.chart import (
    AreaChart,
    BarChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.axis import ChartLines
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.units import EMU_to_cm
from openpyxl.utils.cell import range_boundaries

from .cell_utils import parse_cell_range
from .exceptions import ValidationError, ChartError
from .workbook import _normalize_sheet_reference_name, require_worksheet, safe_workbook

logger = logging.getLogger(__name__)
DEFAULT_CHART_WIDTH = 15.0
DEFAULT_CHART_HEIGHT = 7.5
DEFAULT_COLUMN_WIDTH = 8.43
DEFAULT_ROW_HEIGHT = 15.0

class ChartType(str, Enum):
    """Supported chart types"""
    LINE = "line"
    BAR = "bar"
    PIE = "pie"
    SCATTER = "scatter"
    AREA = "area"

class ChartStyle:
    """Chart style configuration"""
    def __init__(
        self,
        title_size: int = 14,
        title_bold: bool = True,
        axis_label_size: int = 12,
        show_legend: bool = True,
        legend_position: str = "r",
        show_data_labels: bool = True,
        grid_lines: bool = False,
        style_id: int = 2
    ):
        self.title_size = title_size
        self.title_bold = title_bold
        self.axis_label_size = axis_label_size
        self.show_legend = show_legend
        self.legend_position = legend_position
        self.show_data_labels = show_data_labels
        self.grid_lines = grid_lines
        self.style_id = style_id


def _extract_text_runs(rich_text: Any) -> Optional[str]:
    parts: list[str] = []
    for paragraph in getattr(rich_text, "p", []) or []:
        for run in getattr(paragraph, "r", []) or []:
            text = getattr(run, "t", None)
            if text:
                parts.append(text)
        for field in getattr(paragraph, "fld", []) or []:
            text = getattr(field, "t", None)
            if text:
                parts.append(text)
    return "".join(parts) or None


def _extract_title_text(title: Any) -> Optional[str]:
    if title is None:
        return None
    if isinstance(title, str):
        return title or None

    tx = getattr(title, "tx", None)
    if tx is not None:
        rich_text = getattr(tx, "rich", None)
        if rich_text is not None:
            extracted = _extract_text_runs(rich_text)
            if extracted:
                return extracted

        str_ref = getattr(tx, "strRef", None)
        if str_ref is not None and getattr(str_ref, "f", None):
            return str_ref.f

    str_ref = getattr(title, "strRef", None)
    if str_ref is not None and getattr(str_ref, "f", None):
        return str_ref.f

    value = getattr(title, "v", None)
    if value is not None:
        value_text = str(value).strip()
        if not value_text or value_text == "None":
            return None
        return value_text

    return None


def _extract_reference_formula(data_source: Any) -> Optional[str]:
    if data_source is None:
        return None

    for attr_name in ("numRef", "strRef", "multiLvlStrRef"):
        reference = getattr(data_source, attr_name, None)
        if reference is not None and getattr(reference, "f", None):
            return reference.f

    return None


def _extract_chart_anchor(chart: Any) -> Optional[str]:
    anchor = getattr(chart, "anchor", None)
    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None
    return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"


def _extract_series_metadata(series: Any) -> Dict[str, Any]:
    metadata: Dict[str, Any] = {}

    title = _extract_title_text(getattr(series, "tx", None))
    if title:
        metadata["title"] = title

    categories = _extract_reference_formula(getattr(series, "cat", None))
    if categories:
        metadata["categories"] = categories

    x_values = _extract_reference_formula(getattr(series, "xVal", None))
    if x_values:
        metadata["x_values"] = x_values

    values = _extract_reference_formula(getattr(series, "val", None))
    if values:
        metadata["values"] = values

    y_values = _extract_reference_formula(getattr(series, "yVal", None))
    if y_values:
        metadata["y_values"] = y_values

    return metadata


def _extract_chart_dimensions(chart: Any) -> tuple[Optional[float], Optional[float]]:
    anchor = getattr(chart, "anchor", None)
    ext = getattr(anchor, "ext", None)
    if ext is not None and getattr(ext, "cx", None) is not None and getattr(ext, "cy", None) is not None:
        return EMU_to_cm(ext.cx), EMU_to_cm(ext.cy)
    return getattr(chart, "width", None), getattr(chart, "height", None)


def _bounds_to_range(min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _column_width_to_pixels(width: float) -> int:
    if width <= 0:
        return 1
    if width < 1:
        return max(int(round(width * 12)), 1)
    return max(int(math.floor(width * 7 + 5)), 1)


def _row_height_to_pixels(height_points: float) -> int:
    if height_points <= 0:
        return 1
    return max(int(round(height_points * 96 / 72)), 1)


def _cm_to_pixels(value_cm: float) -> int:
    return max(int(math.ceil((value_cm / 2.54) * 96)), 1)


def _column_display_width(worksheet: Any, column_index: int) -> float:
    column_letter = get_column_letter(column_index)
    width = worksheet.column_dimensions[column_letter].width
    if width is not None:
        return float(width)

    default_width = getattr(getattr(worksheet, "sheet_format", None), "defaultColWidth", None)
    if default_width is not None:
        return float(default_width)
    return DEFAULT_COLUMN_WIDTH


def _row_display_height(worksheet: Any, row_index: int) -> float:
    height = worksheet.row_dimensions[row_index].height
    if height is not None:
        return float(height)

    default_height = getattr(getattr(worksheet, "sheet_format", None), "defaultRowHeight", None)
    if default_height is not None:
        return float(default_height)
    return DEFAULT_ROW_HEIGHT


def _chart_bounds_from_anchor(
    worksheet: Any,
    anchor_cell: str,
    *,
    width: float,
    height: float,
) -> tuple[int, int, int, int]:
    _validate_target_cell(anchor_cell)
    start_row, start_col, _, _ = parse_cell_range(anchor_cell)

    remaining_width = _cm_to_pixels(width)
    end_col = start_col
    while remaining_width > 0:
        remaining_width -= _column_width_to_pixels(_column_display_width(worksheet, end_col))
        if remaining_width > 0:
            end_col += 1

    remaining_height = _cm_to_pixels(height)
    end_row = start_row
    while remaining_height > 0:
        remaining_height -= _row_height_to_pixels(_row_display_height(worksheet, end_row))
        if remaining_height > 0:
            end_row += 1

    return start_row, start_col, end_row, end_col


def _grid_bounds(worksheet: Any) -> Optional[tuple[int, int, int, int]]:
    is_empty = (
        worksheet.max_row == 1
        and worksheet.max_column == 1
        and worksheet.cell(1, 1).value is None
    )
    if is_empty:
        return None
    return 1, 1, worksheet.max_row, worksheet.max_column


def _union_bounds(
    first: Optional[tuple[int, int, int, int]],
    second: Optional[tuple[int, int, int, int]],
) -> Optional[tuple[int, int, int, int]]:
    if first is None:
        return second
    if second is None:
        return first

    return (
        min(first[0], second[0]),
        min(first[1], second[1]),
        max(first[2], second[2]),
        max(first[3], second[3]),
    )


def _chart_occupied_range(
    worksheet: Any,
    anchor_cell: str,
    *,
    width: float,
    height: float,
) -> str:
    start_row, start_col, end_row, end_col = _chart_bounds_from_anchor(
        worksheet,
        anchor_cell,
        width=width,
        height=height,
    )
    return _bounds_to_range(start_row, start_col, end_row, end_col)


def _existing_chart_bounds(worksheet: Any) -> Optional[tuple[int, int, int, int]]:
    bounds: Optional[tuple[int, int, int, int]] = None
    for chart in getattr(worksheet, "_charts", []):
        anchor = _extract_chart_anchor(chart)
        if not anchor:
            continue
        width, height = _extract_chart_dimensions(chart)
        chart_bounds = _chart_bounds_from_anchor(
            worksheet,
            anchor,
            width=width or DEFAULT_CHART_WIDTH,
            height=height or DEFAULT_CHART_HEIGHT,
        )
        bounds = _union_bounds(bounds, chart_bounds)
    return bounds


def _worksheet_content_bounds(worksheet: Any) -> Optional[tuple[int, int, int, int]]:
    return _union_bounds(_grid_bounds(worksheet), _existing_chart_bounds(worksheet))


def _bounds_intersect(
    first: tuple[int, int, int, int],
    second: tuple[int, int, int, int],
) -> bool:
    return not (
        first[2] < second[0]
        or second[2] < first[0]
        or first[3] < second[1]
        or second[3] < first[1]
    )


def _inflate_bounds(
    bounds: tuple[int, int, int, int],
    *,
    padding_rows: int = 0,
    padding_columns: int = 0,
) -> tuple[int, int, int, int]:
    return (
        max(bounds[0] - padding_rows, 1),
        max(bounds[1] - padding_columns, 1),
        bounds[2] + padding_rows,
        bounds[3] + padding_columns,
    )


def _occupied_cell_bounds(worksheet: Any) -> list[tuple[int, int, int, int]]:
    raw_cells = getattr(worksheet, "_cells", None)
    if not isinstance(raw_cells, dict) or not raw_cells:
        return []

    row_to_columns: dict[int, list[int]] = {}
    for key, cell in raw_cells.items():
        value = getattr(cell, "value", None)
        if value is None:
            continue

        if isinstance(key, tuple) and len(key) == 2:
            row_index, column_index = key
        else:
            row_index = getattr(cell, "row", None)
            column_index = getattr(cell, "column", None)

        if row_index is None or column_index is None:
            continue
        row_to_columns.setdefault(int(row_index), []).append(int(column_index))

    bounds: list[tuple[int, int, int, int]] = []
    for row_index, columns in row_to_columns.items():
        if not columns:
            continue
        sorted_columns = sorted(set(columns))
        segment_start = sorted_columns[0]
        segment_end = sorted_columns[0]
        for column_index in sorted_columns[1:]:
            if column_index == segment_end + 1:
                segment_end = column_index
                continue
            bounds.append((row_index, segment_start, row_index, segment_end))
            segment_start = column_index
            segment_end = column_index
        bounds.append((row_index, segment_start, row_index, segment_end))

    return bounds


def _merged_range_bounds(worksheet: Any) -> list[tuple[int, int, int, int]]:
    bounds: list[tuple[int, int, int, int]] = []
    for merged_range in getattr(getattr(worksheet, "merged_cells", None), "ranges", []):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        bounds.append((min_row, min_col, max_row, max_col))
    return bounds


def _table_bounds(worksheet: Any) -> list[tuple[int, int, int, int]]:
    if not hasattr(worksheet, "tables"):
        return []

    bounds: list[tuple[int, int, int, int]] = []
    for table in worksheet.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        bounds.append((min_row, min_col, max_row, max_col))
    return bounds


def _existing_chart_occupied_bounds(worksheet: Any) -> list[tuple[int, int, int, int]]:
    bounds: list[tuple[int, int, int, int]] = []
    for chart in getattr(worksheet, "_charts", []):
        anchor = _extract_chart_anchor(chart)
        if not anchor:
            continue
        width, height = _extract_chart_dimensions(chart)
        bounds.append(
            _chart_bounds_from_anchor(
                worksheet,
                anchor,
                width=width or DEFAULT_CHART_WIDTH,
                height=height or DEFAULT_CHART_HEIGHT,
            )
        )
    return bounds


def _occupied_layout_bounds(worksheet: Any) -> dict[str, list[tuple[int, int, int, int]]]:
    return {
        "cells": _occupied_cell_bounds(worksheet),
        "merged_ranges": _merged_range_bounds(worksheet),
        "tables": _table_bounds(worksheet),
        "charts": _existing_chart_occupied_bounds(worksheet),
    }


def _search_window(
    worksheet: Any,
    *,
    search_rows: Optional[int],
    search_columns: Optional[int],
) -> tuple[int, int]:
    occupied_bounds = _occupied_layout_bounds(worksheet)
    max_occupied_row = max(
        (bounds[2] for values in occupied_bounds.values() for bounds in values),
        default=1,
    )
    max_occupied_col = max(
        (bounds[3] for values in occupied_bounds.values() for bounds in values),
        default=1,
    )
    resolved_rows = search_rows if search_rows is not None else max(max_occupied_row + 40, 60)
    resolved_columns = (
        search_columns if search_columns is not None else max(max_occupied_col + 20, 20)
    )
    return resolved_rows, resolved_columns


def _candidate_bounds(
    worksheet: Any,
    anchor_cell: str,
    *,
    width: Optional[float],
    height: Optional[float],
    min_rows: Optional[int],
    min_cols: Optional[int],
) -> tuple[int, int, int, int]:
    if width is not None or height is not None:
        resolved_width, resolved_height = _resolve_chart_dimensions({}, width, height)
        return _chart_bounds_from_anchor(
            worksheet,
            anchor_cell,
            width=resolved_width,
            height=resolved_height,
        )

    if min_rows is None or min_cols is None:
        raise ValidationError("Provide both min_rows and min_cols when width/height are omitted")
    if min_rows <= 0 or min_cols <= 0:
        raise ValidationError("min_rows and min_cols must be positive integers")

    start_row, start_col, _, _ = parse_cell_range(anchor_cell, anchor_cell)
    end_row = start_row + min_rows - 1
    end_col = start_col + min_cols - 1
    return start_row, start_col, end_row, end_col


def _find_free_canvas_slots_in_worksheet(
    worksheet: Any,
    *,
    width: Optional[float] = None,
    height: Optional[float] = None,
    min_rows: Optional[int] = None,
    min_cols: Optional[int] = None,
    limit: int = 5,
    origin_cell: str = "A1",
    search_rows: Optional[int] = None,
    search_columns: Optional[int] = None,
    padding_rows: int = 0,
    padding_columns: int = 0,
) -> dict[str, Any]:
    if limit <= 0:
        raise ValidationError("limit must be a positive integer")
    _validate_target_cell(origin_cell)
    if padding_rows < 0 or padding_columns < 0:
        raise ValidationError("padding_rows and padding_columns must be non-negative integers")

    origin_row, origin_col, _, _ = parse_cell_range(origin_cell, origin_cell)
    max_search_row, max_search_col = _search_window(
        worksheet,
        search_rows=search_rows,
        search_columns=search_columns,
    )
    if origin_row > max_search_row or origin_col > max_search_col:
        raise ValidationError("origin_cell falls outside the requested search window")

    occupied_groups = _occupied_layout_bounds(worksheet)
    blocked_bounds = [
        _inflate_bounds(
            bounds,
            padding_rows=padding_rows,
            padding_columns=padding_columns,
        )
        for values in occupied_groups.values()
        for bounds in values
    ]

    resolved_width: Optional[float] = None
    resolved_height: Optional[float] = None
    if width is not None or height is not None or (min_rows is None and min_cols is None):
        resolved_width, resolved_height = _resolve_chart_dimensions({}, width, height)

    suggestions: list[dict[str, Any]] = []
    reserved_bounds = list(blocked_bounds)
    for row_index in range(origin_row, max_search_row + 1):
        for col_index in range(origin_col, max_search_col + 1):
            anchor_cell = f"{get_column_letter(col_index)}{row_index}"
            candidate = _candidate_bounds(
                worksheet,
                anchor_cell,
                width=resolved_width,
                height=resolved_height,
                min_rows=min_rows,
                min_cols=min_cols,
            )
            if candidate[2] > max_search_row or candidate[3] > max_search_col:
                continue
            if any(_bounds_intersect(candidate, occupied) for occupied in reserved_bounds):
                continue

            suggestion = {
                "anchor_cell": anchor_cell,
                "occupied_range": _bounds_to_range(*candidate),
                "row_span": candidate[2] - candidate[0] + 1,
                "column_span": candidate[3] - candidate[1] + 1,
            }
            if resolved_width is not None and resolved_height is not None:
                suggestion["width"] = resolved_width
                suggestion["height"] = resolved_height
            suggestions.append(suggestion)
            reserved_bounds.append(
                _inflate_bounds(
                    candidate,
                    padding_rows=padding_rows,
                    padding_columns=padding_columns,
                )
            )
            if len(suggestions) >= limit:
                break
        if len(suggestions) >= limit:
            break

    return {
        "sheet_name": worksheet.title,
        "origin_cell": origin_cell,
        "search_window": _bounds_to_range(origin_row, origin_col, max_search_row, max_search_col),
        "occupancy": {
            "cell_segments": len(occupied_groups["cells"]),
            "merged_ranges": len(occupied_groups["merged_ranges"]),
            "tables": len(occupied_groups["tables"]),
            "charts": len(occupied_groups["charts"]),
        },
        "suggestions": suggestions,
    }


def _placement_reference_bounds(
    workbook: Any,
    worksheet: Any,
    *,
    relative_to: Optional[str],
    data_range: Optional[str],
) -> Optional[tuple[int, int, int, int]]:
    reference = relative_to or ""
    normalized_reference = reference.strip().lower()

    if normalized_reference in {"", "content"}:
        return _worksheet_content_bounds(worksheet)

    if normalized_reference == "used_range":
        return _grid_bounds(worksheet)

    if normalized_reference == "data_range":
        if not data_range:
            raise ValidationError("placement.relative_to='data_range' requires data_range")
        source_worksheet, start_row, start_col, end_row, end_col = _resolve_range_source(
            workbook,
            worksheet,
            data_range,
        )
        if source_worksheet.title != worksheet.title:
            raise ValidationError(
                "placement.relative_to='data_range' only works when chart data is on the target worksheet"
            )
        return start_row, start_col, end_row, end_col

    if normalized_reference.startswith("table:"):
        table_name = reference.split(":", 1)[1].strip()
        if not table_name:
            raise ValidationError("placement.relative_to='table:<name>' requires a table name")
        for table in worksheet.tables.values():
            if table.displayName == table_name:
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                return min_row, min_col, max_row, max_col
        raise ValidationError(f"Table '{table_name}' not found in worksheet '{worksheet.title}'")

    if "!" in reference:
        source_worksheet, start_row, start_col, end_row, end_col = _resolve_range_source(
            workbook,
            worksheet,
            reference,
        )
        if source_worksheet.title != worksheet.title:
            raise ValidationError("placement ranges must refer to the target worksheet")
        return start_row, start_col, end_row, end_col

    if ":" in reference:
        try:
            start_cell, end_cell = reference.split(":")
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        except ValueError as exc:
            raise ValidationError(f"Invalid placement range: {reference}") from exc
        return start_row, start_col, end_row, end_col

    raise ValidationError(
        "placement.relative_to must be one of: content, used_range, data_range, "
        "table:<name>, or a worksheet range like A1:C10"
    )


def _resolve_chart_anchor(
    workbook: Any,
    worksheet: Any,
    *,
    target_cell: Optional[str],
    placement: Optional[Dict[str, Any]],
    data_range: Optional[str],
    width: float,
    height: float,
) -> tuple[str, Optional[Dict[str, Any]]]:
    if target_cell and placement:
        raise ValidationError("Provide either target_cell or placement, not both")

    if placement is None:
        if not target_cell:
            raise ValidationError("Either target_cell or placement is required")
        _validate_target_cell(target_cell)
        return target_cell, None

    if not isinstance(placement, dict):
        raise ValidationError("placement must be an object")

    direction = str(placement.get("direction", "right")).strip().lower()
    if direction not in {"right", "below"}:
        raise ValidationError("placement.direction must be either 'right' or 'below'")

    def _coerce_padding(name: str, default: int) -> int:
        raw_value = placement.get(name, default)
        try:
            value = int(raw_value)
        except (TypeError, ValueError) as exc:
            raise ValidationError(f"placement.{name} must be a non-negative integer") from exc
        if value < 0:
            raise ValidationError(f"placement.{name} must be a non-negative integer")
        return value

    padding_columns = _coerce_padding("padding_columns", 1)
    padding_rows = _coerce_padding("padding_rows", 1)
    relative_to = placement.get("relative_to")
    if isinstance(relative_to, str) and relative_to.strip().lower() == "free_canvas":
        search_rows_raw = placement.get("search_rows")
        search_columns_raw = placement.get("search_columns")

        def _coerce_optional_positive_int(raw_value: Any, name: str) -> Optional[int]:
            if raw_value is None:
                return None
            try:
                value = int(raw_value)
            except (TypeError, ValueError) as exc:
                raise ValidationError(f"placement.{name} must be a positive integer") from exc
            if value <= 0:
                raise ValidationError(f"placement.{name} must be a positive integer")
            return value

        search_rows = _coerce_optional_positive_int(search_rows_raw, "search_rows")
        search_columns = _coerce_optional_positive_int(search_columns_raw, "search_columns")
        origin_cell = str(placement.get("origin_cell", "A1")).strip() or "A1"
        free_canvas = _find_free_canvas_slots_in_worksheet(
            worksheet,
            width=width,
            height=height,
            limit=1,
            origin_cell=origin_cell,
            search_rows=search_rows,
            search_columns=search_columns,
            padding_rows=padding_rows,
            padding_columns=padding_columns,
        )
        suggestions = free_canvas["suggestions"]
        if not suggestions:
            raise ValidationError(
                "No free canvas slot found inside the requested search window"
            )
        suggestion = suggestions[0]
        return suggestion["anchor_cell"], {
            "mode": "placement",
            "direction": "free_canvas",
            "relative_to": "free_canvas",
            "padding_columns": padding_columns,
            "padding_rows": padding_rows,
            "occupied_range": suggestion["occupied_range"],
            "search_window": free_canvas["search_window"],
        }

    if relative_to is None:
        resolved_relative_to = "data_range" if data_range else "content"
        try:
            reference_bounds = _placement_reference_bounds(
                workbook,
                worksheet,
                relative_to=resolved_relative_to,
                data_range=data_range,
            )
        except ValidationError:
            resolved_relative_to = "content"
            reference_bounds = _placement_reference_bounds(
                workbook,
                worksheet,
                relative_to=resolved_relative_to,
                data_range=data_range,
            )
    else:
        resolved_relative_to = str(relative_to)
        reference_bounds = _placement_reference_bounds(
            workbook,
            worksheet,
            relative_to=resolved_relative_to,
            data_range=data_range,
        )

    if reference_bounds is None:
        anchor_cell = "A1"
    else:
        min_row, min_col, max_row, max_col = reference_bounds
        if direction == "right":
            anchor_cell = f"{get_column_letter(max_col + padding_columns + 1)}{min_row}"
        else:
            anchor_cell = f"{get_column_letter(min_col)}{max_row + padding_rows + 1}"

    occupied_range = _chart_occupied_range(
        worksheet,
        anchor_cell,
        width=width,
        height=height,
    )
    return anchor_cell, {
        "mode": "placement",
        "direction": direction,
        "relative_to": resolved_relative_to,
        "padding_columns": padding_columns,
        "padding_rows": padding_rows,
        "occupied_range": occupied_range,
    }


def _chart_type_name(chart: Any) -> str:
    class_name = type(chart).__name__
    if class_name.endswith("Chart"):
        return class_name.removesuffix("Chart").lower()
    return class_name.lower()


def list_charts(
    filepath: str,
    sheet_name: Optional[str] = None,
) -> list[dict[str, Any]]:
    """List embedded charts for one worksheet or the whole workbook."""
    try:
        with safe_workbook(filepath) as wb:
            if sheet_name is not None and sheet_name not in wb.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found")

            sheet_names = [sheet_name] if sheet_name is not None else list(wb.sheetnames)
            charts: list[dict[str, Any]] = []

            for current_sheet_name in sheet_names:
                worksheet = wb[current_sheet_name]
                for chart_index, chart in enumerate(getattr(worksheet, "_charts", []), start=1):
                    series = getattr(chart, "ser", None) or list(getattr(chart, "series", []))
                    width, height = _extract_chart_dimensions(chart)
                    anchor = _extract_chart_anchor(chart)
                    chart_info = {
                        "sheet_name": current_sheet_name,
                        "chart_index": chart_index,
                        "chart_type": _chart_type_name(chart),
                        "anchor": anchor,
                        "title": _extract_title_text(getattr(chart, "title", None)),
                        "x_axis_title": _extract_title_text(
                            getattr(getattr(chart, "x_axis", None), "title", None)
                        ),
                        "y_axis_title": _extract_title_text(
                            getattr(getattr(chart, "y_axis", None), "title", None)
                        ),
                        "legend_position": getattr(getattr(chart, "legend", None), "position", None),
                        "style": getattr(chart, "style", None),
                        "width": width,
                        "height": height,
                        "series": [_extract_series_metadata(item) for item in series],
                    }
                    if anchor and width and height:
                        chart_info["occupied_range"] = _chart_occupied_range(
                            worksheet,
                            anchor,
                            width=width,
                            height=height,
                        )
                    charts.append({key: value for key, value in chart_info.items() if value is not None})

            return charts

    except ValidationError:
        raise
    except Exception as e:
        logger.error(f"Failed to list charts: {e}")
        raise ChartError(str(e))


def find_free_canvas_slots(
    filepath: str,
    sheet_name: str,
    *,
    width: Optional[float] = None,
    height: Optional[float] = None,
    min_rows: Optional[int] = None,
    min_cols: Optional[int] = None,
    limit: int = 5,
    origin_cell: str = "A1",
    search_rows: Optional[int] = None,
    search_columns: Optional[int] = None,
    padding_rows: int = 0,
    padding_columns: int = 0,
) -> dict[str, Any]:
    """Suggest free worksheet slots for charts or dashboard blocks."""
    try:
        with safe_workbook(filepath) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=ValidationError,
                operation="layout inspection",
            )
            return _find_free_canvas_slots_in_worksheet(
                worksheet,
                width=width,
                height=height,
                min_rows=min_rows,
                min_cols=min_cols,
                limit=limit,
                origin_cell=origin_cell,
                search_rows=search_rows,
                search_columns=search_columns,
                padding_rows=padding_rows,
                padding_columns=padding_columns,
            )
    except ValidationError:
        raise
    except Exception as e:
        logger.error(f"Failed to find free canvas slots: {e}")
        raise ChartError(str(e))


def _validate_target_cell(target_cell: str) -> None:
    if not target_cell:
        raise ValidationError("Invalid target cell format: target cell is required")

    column_part = "".join(character for character in target_cell if character.isalpha())
    row_part = "".join(character for character in target_cell if character.isdigit())
    if not column_part or not row_part:
        raise ValidationError(f"Invalid target cell format: {target_cell}")

    try:
        column_index_from_string(column_part)
        row_index = int(row_part)
    except ValueError as e:
        raise ValidationError(f"Invalid target cell: {str(e)}") from e

    if row_index < 1:
        raise ValidationError(f"Invalid target cell: {target_cell}")


def _normalize_style(style: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    normalized_style = dict(style or {})
    normalized_style.setdefault("show_data_labels", True)
    return normalized_style


def _resolve_chart_class(chart_type: str) -> tuple[str, Any]:
    chart_classes = {
        "line": LineChart,
        "bar": BarChart,
        "pie": PieChart,
        "scatter": ScatterChart,
        "area": AreaChart,
    }

    chart_type_lower = chart_type.lower()
    chart_class = chart_classes.get(chart_type_lower)
    if not chart_class:
        raise ValidationError(
            f"Unsupported chart type: {chart_type}. "
            f"Supported types: {', '.join(chart_classes.keys())}"
        )
    return chart_type_lower, chart_class


def _build_chart(
    chart_type: str,
    *,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
) -> Any:
    _, chart_class = _resolve_chart_class(chart_type)
    chart = chart_class()
    if title:
        chart.title = title
    if hasattr(chart, "x_axis") and x_axis:
        chart.x_axis.title = x_axis
    if hasattr(chart, "y_axis") and y_axis:
        chart.y_axis.title = y_axis
    return chart


def _resolve_chart_dimensions(
    style: Optional[Dict[str, Any]],
    width: Optional[float],
    height: Optional[float],
) -> tuple[float, float]:
    style = style or {}

    raw_width = width if width is not None else style.get("width")
    raw_height = height if height is not None else style.get("height")

    def _coerce_dimension(raw_value: Any, name: str, default: float) -> float:
        if raw_value is None:
            return default
        try:
            value = float(raw_value)
        except (TypeError, ValueError) as exc:
            raise ValidationError(f"{name} must be a positive number") from exc
        if value <= 0:
            raise ValidationError(f"{name} must be a positive number")
        return value

    return (
        _coerce_dimension(raw_width, "width", DEFAULT_CHART_WIDTH),
        _coerce_dimension(raw_height, "height", DEFAULT_CHART_HEIGHT),
    )


def _resolve_range_source(
    workbook: Any,
    default_worksheet: Any,
    range_ref: str,
) -> tuple[Any, int, int, int, int]:
    if not range_ref:
        raise ValidationError("Range reference is required")

    if "!" in range_ref:
        range_sheet_name, cell_range = range_ref.rsplit("!", 1)
        range_sheet_name = _normalize_sheet_reference_name(range_sheet_name)
        source_worksheet = require_worksheet(
            workbook,
            range_sheet_name,
            error_cls=ValidationError,
            operation="chart data ranges",
        )
    else:
        source_worksheet = default_worksheet
        cell_range = range_ref

    if ":" not in cell_range:
        raise ValidationError(f"Invalid data range format: {range_ref}")

    try:
        start_cell, end_cell = cell_range.split(":")
        start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
    except ValueError as e:
        raise ValidationError(f"Invalid data range format: {str(e)}") from e

    return source_worksheet, start_row, start_col, end_row, end_col


def _validate_contiguous_chart_source(
    *,
    chart_type: str,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
) -> None:
    data_row_count = end_row - start_row
    value_column_count = end_col - start_col

    if data_row_count < 1:
        raise ValidationError(
            "data_range must include a header row and at least one data row"
        )

    if value_column_count < 1:
        if chart_type == "scatter":
            raise ValidationError(
                "Scatter chart data_range must include an X column and at least one Y series column"
            )
        raise ValidationError(
            "data_range must include a category column and at least one value column"
        )


def _reference_from_range(
    workbook: Any,
    default_worksheet: Any,
    range_ref: str,
) -> Reference:
    source_worksheet, start_row, start_col, end_row, end_col = _resolve_range_source(
        workbook,
        default_worksheet,
        range_ref,
    )
    return Reference(
        source_worksheet,
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    )


def _apply_chart_style(chart: Any, style: Dict[str, Any]) -> None:
    try:
        if style.get("show_legend", True):
            chart.legend = Legend()
            chart.legend.position = style.get("legend_position", "r")
        else:
            chart.legend = None

        if style.get("show_data_labels", False):
            data_labels = DataLabelList()
            data_label_options = style.get("data_label_options", {})
            if not isinstance(data_label_options, dict):
                data_label_options = {}

            def _opt(name: str, default: bool) -> bool:
                return bool(data_label_options.get(name, default))

            data_labels.showVal = _opt("show_val", True)
            data_labels.showCatName = _opt("show_cat_name", False)
            data_labels.showSerName = _opt("show_ser_name", False)
            data_labels.showLegendKey = _opt("show_legend_key", False)
            data_labels.showPercent = _opt("show_percent", False)
            data_labels.showBubbleSize = _opt("show_bubble_size", False)
            chart.dataLabels = data_labels

        if style.get("grid_lines", False):
            if hasattr(chart, "x_axis"):
                chart.x_axis.majorGridlines = ChartLines()
            if hasattr(chart, "y_axis"):
                chart.y_axis.majorGridlines = ChartLines()
    except Exception as e:
        logger.error(f"Failed to apply chart style: {e}")
        raise ChartError(f"Failed to apply chart style: {str(e)}")


def _finalize_chart(
    worksheet: Any,
    chart: Any,
    target_cell: str,
    *,
    width: float = DEFAULT_CHART_WIDTH,
    height: float = DEFAULT_CHART_HEIGHT,
) -> None:
    try:
        _validate_target_cell(target_cell)
        chart.width = width
        chart.height = height
        worksheet.add_chart(chart, target_cell)
    except ValidationError:
        raise
    except Exception as e:
        logger.error(f"Failed to create chart drawing: {e}")
        raise ChartError(f"Failed to create chart drawing: {str(e)}")


def create_chart_in_sheet(
    filepath: str,
    sheet_name: str,
    data_range: Optional[str],
    chart_type: str,
    target_cell: Optional[str] = None,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
    style: Optional[Dict] = None,
    series: Optional[List[Dict[str, Any]]] = None,
    categories_range: Optional[str] = None,
    width: Optional[float] = None,
    height: Optional[float] = None,
    placement: Optional[Dict[str, Any]] = None,
) -> dict[str, Any]:
    """Create chart in sheet with either a contiguous data range or explicit series."""
    style = _normalize_style(style)
    resolved_width, resolved_height = _resolve_chart_dimensions(style, width, height)
    if data_range and series:
        raise ValidationError("Provide either data_range or series, not both")
    if not data_range and not series:
        raise ValidationError("Either data_range or series is required")
    if series is not None:
        return create_chart_from_series(
            filepath=filepath,
            sheet_name=sheet_name,
            chart_type=chart_type,
            target_cell=target_cell,
            series=series,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis,
            categories_range=categories_range,
            style=style,
            width=resolved_width,
            height=resolved_height,
            placement=placement,
        )

    try:
        with safe_workbook(filepath, save=True) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=ValidationError,
                operation="creating embedded charts",
            )
            source_worksheet, start_row, start_col, end_row, end_col = _resolve_range_source(
                wb,
                worksheet,
                data_range,
            )
            chart_type_lower, _ = _resolve_chart_class(chart_type)
            _validate_contiguous_chart_source(
                chart_type=chart_type_lower,
                start_row=start_row,
                start_col=start_col,
                end_row=end_row,
                end_col=end_col,
            )
            chart = _build_chart(chart_type_lower, title=title, x_axis=x_axis, y_axis=y_axis)
            resolved_target_cell, placement_details = _resolve_chart_anchor(
                wb,
                worksheet,
                target_cell=target_cell,
                placement=placement,
                data_range=data_range,
                width=resolved_width,
                height=resolved_height,
            )

            try:
                if chart_type_lower == "scatter":
                    for col in range(start_col + 1, end_col + 1):
                        series_title = source_worksheet.cell(row=start_row, column=col).value
                        x_values = Reference(
                            source_worksheet,
                            min_row=start_row + 1,
                            max_row=end_row,
                            min_col=start_col,
                        )
                        y_values = Reference(
                            source_worksheet,
                            min_row=start_row + 1,
                            max_row=end_row,
                            min_col=col,
                        )
                        series = Series(y_values, x_values, title=series_title)
                        chart.series.append(series)
                else:
                    data = Reference(
                        source_worksheet,
                        min_row=start_row,
                        max_row=end_row,
                        min_col=start_col + 1,
                        max_col=end_col,
                    )
                    cats = Reference(
                        source_worksheet,
                        min_row=start_row + 1,
                        max_row=end_row,
                        min_col=start_col,
                    )
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
            except Exception as e:
                logger.error(f"Failed to create chart data references: {e}")
                raise ChartError(f"Failed to create chart data references: {str(e)}")

            _apply_chart_style(chart, style)
            _finalize_chart(
                worksheet,
                chart,
                resolved_target_cell,
                width=resolved_width,
                height=resolved_height,
            )
            occupied_range = _chart_occupied_range(
                worksheet,
                resolved_target_cell,
                width=resolved_width,
                height=resolved_height,
            )

        details = {
            "type": chart_type,
            "location": resolved_target_cell,
            "data_range": data_range,
            "width": resolved_width,
            "height": resolved_height,
            "occupied_range": occupied_range,
        }
        if placement_details is not None:
            details["placement"] = placement_details

        return {
            "message": f"{chart_type.capitalize()} chart created successfully",
            "details": details,
        }

    except (ValidationError, ChartError):
        raise
    except Exception as e:
        logger.error(f"Unexpected error creating chart: {e}")
        raise ChartError(f"Unexpected error creating chart: {str(e)}")


def create_chart_from_series(
    filepath: str,
    sheet_name: str,
    chart_type: str,
    target_cell: Optional[str] = None,
    series: Optional[List[Dict[str, Any]]] = None,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
    categories_range: Optional[str] = None,
    style: Optional[Dict[str, Any]] = None,
    width: Optional[float] = None,
    height: Optional[float] = None,
    placement: Optional[Dict[str, Any]] = None,
) -> dict[str, Any]:
    """Create a chart from explicit series definitions."""
    normalized_style = _normalize_style(style)
    resolved_width, resolved_height = _resolve_chart_dimensions(normalized_style, width, height)
    if not isinstance(series, list) or not series:
        raise ValidationError("At least one series definition is required")

    try:
        with safe_workbook(filepath, save=True) as wb:
            worksheet = require_worksheet(
                wb,
                sheet_name,
                error_cls=ValidationError,
                operation="creating embedded charts",
            )
            chart_type_lower, _ = _resolve_chart_class(chart_type)
            chart = _build_chart(chart_type_lower, title=title, x_axis=x_axis, y_axis=y_axis)
            resolved_target_cell, placement_details = _resolve_chart_anchor(
                wb,
                worksheet,
                target_cell=target_cell,
                placement=placement,
                data_range=None,
                width=resolved_width,
                height=resolved_height,
            )

            if chart_type_lower == "scatter" and categories_range is not None:
                raise ValidationError("categories_range is not supported for scatter charts")
            if chart_type_lower == "pie" and len(series) != 1:
                raise ValidationError("Pie charts require exactly one series definition")

            shared_categories = None
            if chart_type_lower != "scatter" and categories_range is not None:
                shared_categories = _reference_from_range(wb, worksheet, categories_range)

            try:
                for index, series_definition in enumerate(series, start=1):
                    if not isinstance(series_definition, dict):
                        raise ValidationError("Each series definition must be an object")

                    series_title = series_definition.get("title")
                    if chart_type_lower == "scatter":
                        x_range = series_definition.get("x_range")
                        y_range = series_definition.get("y_range")
                        if not x_range or not y_range:
                            raise ValidationError(
                                f"Scatter series {index} requires both x_range and y_range"
                            )

                        x_values = _reference_from_range(wb, worksheet, x_range)
                        y_values = _reference_from_range(wb, worksheet, y_range)
                        chart.series.append(Series(y_values, x_values, title=series_title))
                        continue

                    values_range = series_definition.get("values_range")
                    if not values_range:
                        raise ValidationError(
                            f"Series {index} requires values_range for {chart_type_lower} charts"
                        )

                    values = _reference_from_range(wb, worksheet, values_range)
                    chart.series.append(Series(values, title=series_title))

                if shared_categories is not None:
                    chart.set_categories(shared_categories)
            except ValidationError:
                raise
            except Exception as e:
                logger.error(f"Failed to create chart series: {e}")
                raise ChartError(f"Failed to create chart series: {str(e)}")

            _apply_chart_style(chart, normalized_style)
            _finalize_chart(
                worksheet,
                chart,
                resolved_target_cell,
                width=resolved_width,
                height=resolved_height,
            )
            occupied_range = _chart_occupied_range(
                worksheet,
                resolved_target_cell,
                width=resolved_width,
                height=resolved_height,
            )

        details = {
            "type": chart_type,
            "location": resolved_target_cell,
            "series_count": len(series),
            "categories_range": categories_range,
            "width": resolved_width,
            "height": resolved_height,
            "occupied_range": occupied_range,
        }
        if placement_details is not None:
            details["placement"] = placement_details

        return {
            "message": f"{chart_type.capitalize()} chart created successfully",
            "details": details,
        }

    except (ValidationError, ChartError):
        raise
    except Exception as e:
        logger.error(f"Unexpected error creating chart from series: {e}")
        raise ChartError(f"Unexpected error creating chart from series: {str(e)}")
