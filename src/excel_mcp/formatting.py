import logging
from typing import Any, Dict, List, Optional

from openpyxl.styles import (
    PatternFill, Border, Side, Alignment, Protection, Font,
    Color
)
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
    FormulaRule, CellIsRule
)
from openpyxl.worksheet.worksheet import Worksheet

from .workbook import safe_workbook
from .cell_utils import parse_cell_range, validate_cell_reference
from .exceptions import ValidationError, FormattingError

logger = logging.getLogger(__name__)


def _should_include_changes(dry_run: bool, include_changes: Optional[bool]) -> bool:
    if include_changes is None:
        return dry_run
    return include_changes


def _build_format_preview(
    *,
    sheet_name: str,
    range_str: str,
    bold: bool,
    italic: bool,
    underline: bool,
    font_size: Optional[int],
    font_color: Optional[str],
    bg_color: Optional[str],
    border_style: Optional[str],
    border_color: Optional[str],
    number_format: Optional[str],
    alignment: Optional[str],
    wrap_text: bool,
    merge_cells: bool,
    protection: Optional[Dict[str, Any]],
    conditional_format: Optional[Dict[str, Any]],
) -> Dict[str, Any]:
    return {
        "sheet_name": sheet_name,
        "range": range_str,
        "font": {
            "bold": bold,
            "italic": italic,
            "underline": underline,
            "font_size": font_size,
            "font_color": font_color,
        },
        "fill": {"bg_color": bg_color} if bg_color is not None else None,
        "border": {
            "border_style": border_style,
            "border_color": border_color,
        }
        if border_style is not None or border_color is not None
        else None,
        "number_format": number_format,
        "alignment": alignment,
        "wrap_text": wrap_text,
        "merge_cells": merge_cells,
        "protection": protection,
        "conditional_format": conditional_format,
    }


def _apply_conditional_format(
    sheet: Worksheet,
    range_str: str,
    conditional_format: Dict[str, Any],
) -> None:
    rule_type = conditional_format.get("type")
    if not rule_type:
        raise FormattingError("Conditional format type not specified")

    nested_params = conditional_format.get("params", {})
    if nested_params is None:
        nested_params = {}
    if not isinstance(nested_params, dict):
        raise FormattingError("Conditional format params must be an object")

    top_level_params = {
        key: value
        for key, value in conditional_format.items()
        if key not in {"type", "params"}
    }
    params = dict(top_level_params)
    params.update(nested_params)

    if rule_type == "cell_is" and "fill" in params:
        fill_params = params["fill"]
        if isinstance(fill_params, dict):
            try:
                fill_color = fill_params.get("fgColor", "FFC7CE")
                fill_color = (
                    fill_color if fill_color.startswith("FF") else f"FF{fill_color}"
                )
                params["fill"] = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid",
                )
            except ValueError as e:
                raise FormattingError(
                    f"Invalid conditional format fill color: {str(e)}"
                ) from e

    try:
        if rule_type == "color_scale":
            rule = ColorScaleRule(**params)
        elif rule_type == "data_bar":
            rule = DataBarRule(**params)
        elif rule_type == "icon_set":
            rule = IconSetRule(**params)
        elif rule_type == "formula":
            rule = FormulaRule(**params)
        elif rule_type == "cell_is":
            rule = CellIsRule(**params)
        else:
            raise FormattingError(f"Invalid conditional format type: {rule_type}")

        sheet.conditional_formatting.add(range_str, rule)
    except FormattingError:
        raise
    except Exception as e:
        hint = ""
        if not params:
            hint = (
                " Provide rule parameters under conditional_format.params "
                "or as top-level keys."
            )
        raise FormattingError(
            f"Failed to apply conditional formatting: {str(e)}{hint}"
        ) from e


def _apply_format_to_sheet(
    sheet: Worksheet,
    *,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    if not validate_cell_reference(start_cell):
        raise ValidationError(f"Invalid start cell reference: {start_cell}")

    if end_cell and not validate_cell_reference(end_cell):
        raise ValidationError(f"Invalid end cell reference: {end_cell}")

    try:
        start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
    except ValueError as e:
        raise ValidationError(f"Invalid cell range: {str(e)}") from e

    if end_row is None:
        end_row = start_row
    if end_col is None:
        end_col = start_col

    range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
    preview = _build_format_preview(
        sheet_name=sheet_name,
        range_str=range_str,
        bold=bold,
        italic=italic,
        underline=underline,
        font_size=font_size,
        font_color=font_color,
        bg_color=bg_color,
        border_style=border_style,
        border_color=border_color,
        number_format=number_format,
        alignment=alignment,
        wrap_text=wrap_text,
        merge_cells=merge_cells,
        protection=protection,
        conditional_format=conditional_format,
    )

    font_args = {
        "bold": bold,
        "italic": italic,
        "underline": "single" if underline else None,
    }
    if font_size is not None:
        font_args["size"] = font_size
    if font_color is not None:
        try:
            font_color = (
                font_color if font_color.startswith("FF") else f"FF{font_color}"
            )
            font_args["color"] = Color(rgb=font_color)
        except ValueError as e:
            raise FormattingError(f"Invalid font color: {str(e)}") from e
    font = Font(**font_args)

    fill = None
    if bg_color is not None:
        try:
            bg_color = bg_color if bg_color.startswith("FF") else f"FF{bg_color}"
            fill = PatternFill(
                start_color=Color(rgb=bg_color),
                end_color=Color(rgb=bg_color),
                fill_type="solid",
            )
        except ValueError as e:
            raise FormattingError(f"Invalid background color: {str(e)}") from e

    border = None
    if border_style is not None:
        try:
            normalized_border_color = border_color if border_color else "000000"
            normalized_border_color = (
                normalized_border_color
                if normalized_border_color.startswith("FF")
                else f"FF{normalized_border_color}"
            )
            side = Side(style=border_style, color=Color(rgb=normalized_border_color))
            border = Border(left=side, right=side, top=side, bottom=side)
        except ValueError as e:
            raise FormattingError(f"Invalid border settings: {str(e)}") from e

    align = None
    if alignment is not None or wrap_text:
        try:
            align = Alignment(
                horizontal=alignment,
                vertical="center",
                wrap_text=wrap_text,
            )
        except ValueError as e:
            raise FormattingError(f"Invalid alignment settings: {str(e)}") from e

    protect = None
    if protection is not None:
        try:
            protect = Protection(**protection)
        except ValueError as e:
            raise FormattingError(f"Invalid protection settings: {str(e)}") from e

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = font
            if fill is not None:
                cell.fill = fill
            if border is not None:
                cell.border = border
            if align is not None:
                cell.alignment = align
            if protect is not None:
                cell.protection = protect
            if number_format is not None:
                cell.number_format = number_format

    if merge_cells and end_cell:
        try:
            sheet.merge_cells(range_str)
        except ValueError as e:
            raise FormattingError(f"Failed to merge cells: {str(e)}") from e

    if conditional_format is not None:
        _apply_conditional_format(sheet, range_str, conditional_format)

    return {"range": range_str, "preview": preview}

def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None,
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Apply formatting to a range of cells.
    
    This function handles all Excel formatting operations including:
    - Font properties (bold, italic, size, color, etc.)
    - Cell fill/background color
    - Borders (style and color)
    - Number formatting
    - Alignment and text wrapping
    - Cell merging
    - Protection
    - Conditional formatting
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell reference
        end_cell: Optional ending cell reference
        bold: Whether to make text bold
        italic: Whether to make text italic
        underline: Whether to underline text
        font_size: Font size in points
        font_color: Font color (hex code)
        bg_color: Background color (hex code)
        border_style: Border style (thin, medium, thick, double)
        border_color: Border color (hex code)
        number_format: Excel number format string
        alignment: Text alignment (left, center, right, justify)
        wrap_text: Whether to wrap text
        merge_cells: Whether to merge the range
        protection: Cell protection settings
        conditional_format: Conditional formatting rules
        
    Returns:
        Dictionary with operation status
    """
    try:
        with safe_workbook(filepath, save=not dry_run) as wb:
            if sheet_name not in wb.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found")

            sheet = wb[sheet_name]
            applied = _apply_format_to_sheet(
                sheet,
                sheet_name=sheet_name,
                start_cell=start_cell,
                end_cell=end_cell,
                bold=bold,
                italic=italic,
                underline=underline,
                font_size=font_size,
                font_color=font_color,
                bg_color=bg_color,
                border_style=border_style,
                border_color=border_color,
                number_format=number_format,
                alignment=alignment,
                wrap_text=wrap_text,
                merge_cells=merge_cells,
                protection=protection,
                conditional_format=conditional_format,
            )

        result = {
            "message": f"{'Previewed' if dry_run else 'Applied'} formatting to range {applied['range']}",
            "range": applied["range"],
            "dry_run": dry_run,
        }
        if _should_include_changes(dry_run, include_changes):
            result["changes"] = [applied["preview"]]
        return result

    except (ValidationError, FormattingError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to apply formatting: {e}")
        raise FormattingError(str(e))


def format_ranges(
    filepath: str,
    sheet_name: str,
    ranges: List[Dict[str, Any]],
    dry_run: bool = False,
    include_changes: Optional[bool] = None,
) -> Dict[str, Any]:
    """Apply formatting to multiple ranges in a single workbook pass."""
    try:
        if not ranges:
            raise FormattingError("At least one format operation must be provided")

        with safe_workbook(filepath, save=not dry_run) as wb:
            if sheet_name not in wb.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found")

            sheet = wb[sheet_name]
            applied_ranges: List[str] = []
            previews: List[Dict[str, Any]] = []
            errors: List[Dict[str, Any]] = []

            for index, operation in enumerate(ranges, start=1):
                if not isinstance(operation, dict):
                    raise FormattingError(
                        f"Format operation at index {index} must be an object"
                    )

                operation_data = dict(operation)
                operation_data.pop("dry_run", None)
                operation_data.pop("include_changes", None)
                range_label = operation_data.get("start_cell")
                if operation_data.get("end_cell"):
                    range_label = f"{range_label}:{operation_data['end_cell']}"

                try:
                    applied = _apply_format_to_sheet(
                        sheet,
                        sheet_name=sheet_name,
                        **operation_data,
                    )
                except (ValidationError, FormattingError) as e:
                    errors.append(
                        {
                            "index": index,
                            "range": range_label,
                            "error": str(e),
                        }
                    )
                    continue
                except TypeError as e:
                    errors.append(
                        {
                            "index": index,
                            "range": range_label,
                            "error": f"Invalid format operation at index {index}: {str(e)}",
                        }
                    )
                    continue

                applied_ranges.append(applied["range"])
                previews.append(applied["preview"])

        ranges_failed = len(errors)
        if applied_ranges and not errors:
            message = (
                f"{'Previewed' if dry_run else 'Applied'} formatting to "
                f"{len(applied_ranges)} range(s) in sheet '{sheet_name}'"
            )
        elif applied_ranges:
            message = (
                f"{'Previewed' if dry_run else 'Applied'} formatting to "
                f"{len(applied_ranges)} range(s) in sheet '{sheet_name}'; "
                f"{ranges_failed} range(s) failed"
            )
        else:
            message = (
                f"No formatting operations were applied in sheet '{sheet_name}'; "
                f"{ranges_failed} range(s) failed"
            )

        result = {
            "message": message,
            "sheet_name": sheet_name,
            "ranges_formatted": len(applied_ranges),
            "ranges_failed": ranges_failed,
            "ranges": applied_ranges,
            "dry_run": dry_run,
        }
        if errors:
            result["errors"] = errors
            result["warnings"] = [
                f"{ranges_failed} range(s) failed during batch formatting"
            ]
        if _should_include_changes(dry_run, include_changes):
            result["changes"] = previews
        return result

    except (ValidationError, FormattingError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to apply batch formatting: {e}")
        raise FormattingError(str(e))
