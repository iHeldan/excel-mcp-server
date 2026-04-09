from typing import Any
import uuid
import logging

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

from .data import read_excel_range
from .cell_utils import parse_cell_range
from .exceptions import ValidationError, PivotError
from .workbook import safe_workbook

logger = logging.getLogger(__name__)


def _clean_field_name(field: str) -> str:
    field = str(field).strip()
    for suffix in [" (sum)", " (average)", " (count)", " (min)", " (max)"]:
        if field.lower().endswith(suffix):
            return field[:-len(suffix)]
    return field


def _field_lookup_key(field: str) -> str:
    return _clean_field_name(field).lower()


def _resolve_field_names(requested_fields: list[str], available_fields: list[str], field_type: str) -> list[str]:
    available_lookup: dict[str, str] = {}
    for field in available_fields:
        lookup_key = _field_lookup_key(field)
        if lookup_key in available_lookup and available_lookup[lookup_key] != field:
            raise ValidationError(
                f"Ambiguous {field_type} field '{field}'. Available fields conflict after normalization."
            )
        available_lookup[lookup_key] = field

    resolved_fields: list[str] = []
    for field in requested_fields:
        lookup_key = _field_lookup_key(str(field))
        resolved_field = available_lookup.get(lookup_key)
        if resolved_field is None:
            raise ValidationError(
                f"Invalid {field_type} field '{field}'. "
                f"Available fields: {', '.join(sorted(available_fields))}"
            )
        resolved_fields.append(resolved_field)

    return resolved_fields


def _format_column_header(
    column_filters: dict[str, Any],
    value_field: str,
    agg_func: str,
    *,
    include_value_field: bool,
) -> str:
    value_label = f"{value_field} ({agg_func})"
    if not column_filters:
        return value_label

    if len(column_filters) == 1:
        column_label = str(next(iter(column_filters.values())))
    else:
        column_label = " | ".join(f"{field}={value}" for field, value in column_filters.items())

    if include_value_field:
        return f"{column_label} | {value_label}"
    return column_label

def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: list[str],
    values: list[str],
    columns: list[str] | None = None,
    agg_func: str = "sum"
) -> dict[str, Any]:
    """Create pivot table in sheet using Excel table functionality
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet containing source data
        data_range: Source data range reference
        target_cell: Cell reference for pivot table position
        rows: Fields for row labels
        values: Fields for values
        columns: Optional fields for column labels
        agg_func: Aggregation function (sum, count, average, max, min)
        
    Returns:
        Dictionary with status message and pivot table dimensions
    """
    try:
        # Parse ranges first (no workbook needed)
        if ':' not in data_range:
            raise ValidationError("Data range must be in format 'A1:B2'")

        try:
            start_cell, end_cell = data_range.split(':')
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        except ValueError as e:
            raise ValidationError(f"Invalid data range format: {str(e)}")

        if end_row is None or end_col is None:
            raise ValidationError("Invalid data range format: missing end coordinates")

        # Create range string
        data_range_str = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

        # Read source data and convert to list of dicts
        try:
            data_as_list = read_excel_range(filepath, sheet_name, start_cell, end_cell)
            if not data_as_list or len(data_as_list) < 2:
                raise PivotError("Source data must have a header row and at least one data row.")

            headers = [str(h) for h in data_as_list[0]]
            data = [dict(zip(headers, row)) for row in data_as_list[1:]]

            if not data:
                raise PivotError("No data rows found after header.")

        except Exception as e:
            raise PivotError(f"Failed to read or process source data: {str(e)}")

        # Validate aggregation function
        valid_agg_funcs = ["sum", "average", "count", "min", "max"]
        if agg_func.lower() not in valid_agg_funcs:
            raise ValidationError(
                f"Invalid aggregation function. Must be one of: {', '.join(valid_agg_funcs)}"
            )

        # Validate field names exist in data
        if data:
            available_fields_raw = list(data[0].keys())
            resolved_rows = _resolve_field_names(rows, available_fields_raw, "row")
            resolved_values = _resolve_field_names(values, available_fields_raw, "value")
            resolved_columns = _resolve_field_names(columns or [], available_fields_raw, "column")
        else:
            resolved_rows = rows
            resolved_values = values
            resolved_columns = columns or []

        with safe_workbook(filepath, save=True) as wb:
            if sheet_name not in wb.sheetnames:
                raise ValidationError(f"Sheet '{sheet_name}' not found")

            # Create pivot sheet
            pivot_sheet_name = f"{sheet_name}_pivot"
            if pivot_sheet_name in wb.sheetnames:
                wb.remove(wb[pivot_sheet_name])
            pivot_ws = wb.create_sheet(pivot_sheet_name)

            # Write headers
            current_row = 1
            current_col = 1

            # Write row field headers
            for field in resolved_rows:
                cell = pivot_ws.cell(row=current_row, column=current_col, value=field)
                cell.font = Font(bold=True)
                current_col += 1

            # Resolve row and column combinations before writing value headers
            row_field_values = {field: {record.get(field) for record in data} for field in resolved_rows}
            row_combinations = _get_combinations(row_field_values)

            column_field_values = {field: {record.get(field) for record in data} for field in resolved_columns}
            column_combinations = _get_combinations(column_field_values)

            if not column_combinations:
                column_combinations = [{}]

            include_value_field = len(resolved_values) > 1 or len(resolved_columns) > 1

            # Write value/column headers
            for column_filters in column_combinations:
                for field in resolved_values:
                    header = _format_column_header(
                        column_filters,
                        field,
                        agg_func,
                        include_value_field=include_value_field,
                    )
                    cell = pivot_ws.cell(row=current_row, column=current_col, value=header)
                    cell.font = Font(bold=True)
                    current_col += 1

            if not resolved_values:
                raise ValidationError("At least one value field is required")

            # Recalculate after header writing
            current_row = 2

            # Get unique values for each row field
            field_values = row_field_values

            # Generate all combinations of row field values
            row_combinations = _get_combinations(field_values)

            # Calculate table dimensions for formatting
            total_rows = len(row_combinations) + 1  # +1 for header
            total_cols = len(resolved_rows) + (len(column_combinations) * len(resolved_values))

            # Write data rows
            for combo in row_combinations:
                # Write row field values
                col = 1
                for field in resolved_rows:
                    pivot_ws.cell(row=current_row, column=col, value=combo[field])
                    col += 1

                # Calculate and write aggregated values for each column combination
                for column_filters in column_combinations:
                    filtered_data = _filter_data(data, combo, column_filters)
                    for value_field in resolved_values:
                        try:
                            value = _aggregate_values(filtered_data, value_field, agg_func)
                            pivot_ws.cell(row=current_row, column=col, value=value)
                        except Exception as e:
                            raise PivotError(
                                f"Failed to aggregate values for field '{value_field}': {str(e)}"
                            )
                        col += 1

                current_row += 1

            # Create a table for the pivot data
            try:
                pivot_range = f"A1:{get_column_letter(total_cols)}{total_rows}"
                pivot_table = Table(
                    displayName=f"PivotTable_{uuid.uuid4().hex[:8]}",
                    ref=pivot_range
                )
                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=True
                )
                pivot_table.tableStyleInfo = style
                pivot_ws.add_table(pivot_table)
            except Exception as e:
                raise PivotError(f"Failed to create pivot table formatting: {str(e)}")

        return {
            "message": "Summary table created successfully",
            "details": {
                "source_range": data_range_str,
                "pivot_sheet": pivot_sheet_name,
                "rows": resolved_rows,
                "columns": resolved_columns,
                "values": resolved_values,
                "aggregation": agg_func
            }
        }

    except (ValidationError, PivotError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to create pivot table: {e}")
        raise PivotError(str(e))


def _get_combinations(field_values: dict[str, set]) -> list[dict]:
    """Get all combinations of field values."""
    result = [{}]
    for field, values in list(field_values.items()):  # Convert to list to avoid runtime changes
        new_result = []
        for combo in result:
            for value in sorted(values, key=lambda item: (item is None, str(item))):
                new_combo = combo.copy()
                new_combo[field] = value
                new_result.append(new_combo)
        result = new_result
    return result


def _filter_data(data: list[dict], row_filters: dict, col_filters: dict) -> list[dict]:
    """Filter data based on row and column filters."""
    result = []
    for record in data:
        matches = True
        for field, value in row_filters.items():
            if record.get(field) != value:
                matches = False
                break
        for field, value in col_filters.items():
            if record.get(field) != value:
                matches = False
                break
        if matches:
            result.append(record)
    return result


def _aggregate_values(data: list[dict], field: str, agg_func: str) -> float:
    """Aggregate values using the specified function."""
    values = [record[field] for record in data if field in record and isinstance(record[field], (int, float))]
    if not values:
        return 0
        
    if agg_func == "sum":
        return sum(values)
    elif agg_func == "average":
        return sum(values) / len(values)
    elif agg_func == "count":
        return len(values)
    elif agg_func == "min":
        return min(values)
    elif agg_func == "max":
        return max(values)
    else:
        return sum(values)  # Default to sum
